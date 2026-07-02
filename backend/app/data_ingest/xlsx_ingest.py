# data_ingest/xlsx_ingest.py
# Parses Excel spreadsheets (.xlsx) using openpyxl.

import logging
from typing import List

logger = logging.getLogger(__name__)

def load_documents_from_xlsx(file_path: str) -> List[str]:
    """
    Extracts text representations of rows from all sheets in an Excel workbook.
    """
    rows_text = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            logger.debug(f"Parsing sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                # Format cell values into a string line
                row_values = []
                for val in row:
                    if val is not None:
                        # Strip strings to make output clean
                        row_values.append(str(val).strip())
                
                if row_values:
                    # Represent the row as a pipe-separated line
                    rows_text.append(f"Sheet: {sheet_name} | " + " | ".join(row_values))
                    
        logger.info(f"Extracted {len(rows_text)} rows from Excel workbook: {file_path}")
    except Exception as exc:
        logger.error(f"Error parsing Excel file {file_path}: {exc}")
        raise
        
    return rows_text
